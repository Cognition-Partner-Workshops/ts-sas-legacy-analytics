from datetime import date
from pathlib import Path

import openpyxl
import pytest

from jobs import monthly_regulatory_reporting as mrr
from recon import rules, run_recon
from recon.units import classify_column, tables_for
from sas_macros import export_xlsx


def test_parse_report_month() -> None:
    assert mrr.parse_report_month("202401") == ("202401", date(2024, 1, 31))
    assert mrr.parse_report_month("202402") == ("202402", date(2024, 2, 29))
    for value in ("2024-01", "20241"):
        with pytest.raises(ValueError):
            mrr.parse_report_month(value)


def test_sql_builders_preserve_source_joins_and_filter() -> None:
    sqls = [
        mrr.rwa_sql("202401", date(2024, 1, 31)),
        mrr.delinquency_sql("202401", date(2024, 1, 31)),
        mrr.llp_sql("202401", date(2024, 1, 31)),
        mrr.capital_sql("202401"),
    ]
    assert "LEFT JOIN" in sqls[0]
    assert "sas_bronze.loan_details" in sqls[0]
    assert "snapshot_date = DATE'2024-01-31'" in sqls[0]
    assert "ELSE 1.00" in sqls[0]
    assert "l.ltv <= 0.80" in sqls[0]
    assert "l.ltv > 0.80" in sqls[0]
    assert "REPLACE WHERE report_month = '202401'" in sqls[0]
    assert "'202401' AS report_month" in sqls[0]
    assert "GROUP BY account_type, customer_segment, risk_weight" in sqls[0]
    assert all("daily_transactions" not in sql and "collateral" not in sql for sql in sqls)


def test_delinquency_sql_has_all_buckets_and_loan_types() -> None:
    sql = mrr.delinquency_sql("202401", date(2024, 1, 31))
    for label in ("Current", "1-29", "30-59", "60-89", "90-119", "120-179", "180+", "Unknown"):
        assert f"'{label}'" in sql
    for account_type in ("MTG", "AUTO", "PERS", "CC", "LOC", "HELC"):
        assert f"'{account_type}'" in sql
    assert "l.days_past_due" in sql


def test_llp_sql_uses_inner_join_and_literal_zero_denominators() -> None:
    sql = mrr.llp_sql("202401", date(2024, 1, 31))
    assert "JOIN" in sql and "LEFT JOIN" not in sql
    assert "days_past_due >= 90" in sql
    assert sql.count("ELSE 0") >= 3
    assert "* 100" in sql


def test_capital_sql_has_constants_thresholds_and_zero_status() -> None:
    sql = mrr.capital_sql("202401")
    for value in ("50000000", "65000000", "80000000", "4.5", "6.0", "8.0"):
        assert value in sql
    assert "CAST(NULL AS DOUBLE)" in sql
    assert "sas_gold.monthly_rwa" in sql
    assert "report_month = '202401'" in sql
    assert "total_rwa = 0 THEN 'PASS'" in sql


def test_run_order_sheet_names_and_overwrites_existing_workbook(tmp_path: Path) -> None:
    executed: list[str] = []
    fetched: list[str] = []
    written: list[tuple[str, bool]] = []
    workbook_path = tmp_path / "report.xlsx"
    workbook_path.write_text("old", encoding="utf-8")

    def execute(sql: str) -> None:
        executed.append(sql)

    def fetch(sql: str) -> list[dict[str, object]]:
        fetched.append(sql)
        return [{"report_month": "202401"}]

    def write_sheet(rows, path: str, sheet: str) -> str:
        del rows
        written.append((sheet, Path(path).exists()))
        Path(path).touch()
        return path

    summary = mrr.run(execute, fetch, "202401", date(2024, 1, 31), str(workbook_path), write_sheet)
    assert len(executed) == 8
    assert executed[:4] == mrr.ddl()
    assert "monthly_rwa" in executed[4]
    assert "delinquency_aging" in executed[5]
    assert "llp_coverage" in executed[6]
    assert "capital_adequacy" in executed[7]
    assert [name for name, _exists in written] == ["RWA", "Delinquency", "LLP_Coverage"]
    assert written[0] == ("RWA", False)
    assert "CAPITAL_ADEQUACY" not in fetched[-1].upper() if fetched else True
    assert summary["statements_run"] == 8
    assert summary["sheets_written"] == ["RWA", "Delinquency", "LLP_Coverage"]


def test_export_xlsx_accumulates_and_replaces_sheets(tmp_path: Path) -> None:
    path = tmp_path / "report.xlsx"
    export_xlsx([{"value": "rwa"}], str(path), "RWA")
    export_xlsx([{"value": "delinq"}], str(path), "Delinquency")
    export_xlsx([{"value": "llp"}], str(path), "LLP_Coverage")
    workbook = openpyxl.load_workbook(path, read_only=True)
    assert workbook.sheetnames == ["RWA", "Delinquency", "LLP_Coverage"]
    workbook.close()

    export_xlsx([{"value": "updated"}], str(path), "RWA")
    workbook = openpyxl.load_workbook(path, read_only=True)
    assert workbook.sheetnames == ["Delinquency", "LLP_Coverage", "RWA"]
    assert list(workbook["RWA"].values) == [("value",), ("updated",)]
    workbook.close()


def test_u4_specs_and_t12_sheet_validation(tmp_path: Path) -> None:
    monthly, delinquency, llp, capital = tables_for("U4")
    assert monthly.keys == ("report_month", "account_type", "customer_segment", "risk_weight")
    assert monthly.xlsx_sheets == ("RWA", "Delinquency", "LLP_Coverage")
    assert classify_column(monthly, "n_accounts") == "T-3"
    assert classify_column(monthly, "total_exposure") == "T-4"
    assert classify_column(monthly, "rwa") == "T-4"
    assert classify_column(delinquency, "n_accounts") == "T-3"
    assert classify_column(delinquency, "total_balance") == "T-4"
    assert classify_column(delinquency, "total_past_due") == "T-4"
    assert classify_column(llp, "n_loans") == "T-3"
    assert classify_column(llp, "gross_loans") == "T-4"
    assert classify_column(llp, "total_allowance") == "T-4"
    assert classify_column(llp, "coverage_pct") == "T-5"
    assert classify_column(llp, "npl_balance") == "T-4"
    assert classify_column(llp, "npl_coverage_pct") == "T-5"
    for column in (
        "total_rwa", "cet1_capital", "tier1_capital", "total_capital",
    ):
        assert classify_column(capital, column) == "T-4"
    for column in ("cet1_ratio", "tier1_ratio", "total_capital_ratio"):
        assert classify_column(capital, column) == "T-5"
    for column in ("cet1_status", "tier1_status", "total_capital_status"):
        assert classify_column(capital, column) == "T-3"

    path = tmp_path / "three.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.title = "RWA"
    workbook.create_sheet("Delinquency")
    workbook.create_sheet("LLP_Coverage")
    workbook.save(path)
    workbook.close()
    assert rules.t12(monthly, str(path)).verdict == "PASS"
    workbook = openpyxl.load_workbook(path)
    workbook.create_sheet("Extra")
    workbook.save(path)
    workbook.close()
    assert rules.t12(monthly, str(path)).verdict == "FAIL"
    workbook = openpyxl.load_workbook(path)
    del workbook["Extra"]
    workbook["Delinquency"].title = "Wrong"
    workbook.save(path)
    workbook.close()
    assert rules.t12(monthly, str(path)).verdict == "FAIL"


def test_u4_fixture_recon_self_check(tmp_path: Path) -> None:
    reference = Path(__file__).parents[2] / "docs" / "migration" / "recon" / "reference"
    assert run_recon.main(
        [
            "--unit", "U4", "--mode", "fixture", "--business-date", "2024-01-31",
            "--out", str(tmp_path), "--target-dir", str(reference),
        ]
    ) == 0
    report = __import__("json").loads((tmp_path / "recon.json").read_text(encoding="utf-8"))
    assert all(
        all(result["verdict"] != "FAIL" for result in table["results"])
        for table in report["tables"]
    )


def test_main_missing_report_month_is_nonzero() -> None:
    with pytest.raises(SystemExit) as exc:
        mrr.main(["--business-date", "2024-01-31"])
    assert exc.value.code != 0
