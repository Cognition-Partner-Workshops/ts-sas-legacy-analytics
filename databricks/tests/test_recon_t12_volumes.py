import hashlib
import io
import json
from pathlib import Path

import openpyxl

from recon import rules, run_recon
from recon.units import tables_for


def _workbook_bytes() -> bytes:
    workbook = openpyxl.Workbook()
    workbook.active.title = "RWA"
    workbook.create_sheet("Delinquency")
    workbook.create_sheet("LLP_Coverage")
    stream = io.BytesIO()
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()


def test_resolve_xlsx_volumes_fetches_to_local_file() -> None:
    content = _workbook_bytes()
    local, info = rules.resolve_xlsx("/Volumes/x/y/z.xlsx", lambda _path: content)
    assert local is not None and Path(local).is_file()
    assert info["xlsx_source_path"] == "/Volumes/x/y/z.xlsx"
    assert info["xlsx_sha256"] == hashlib.sha256(content).hexdigest()
    assert len(info["xlsx_sha256"]) == 64
    assert rules.t12(tables_for("U4")[0], local).verdict == "PASS"


def test_resolve_xlsx_local_path_and_missing_fetch(tmp_path: Path) -> None:
    path = tmp_path / "report.xlsx"
    path.write_bytes(_workbook_bytes())
    local, info = rules.resolve_xlsx(str(path), None)
    assert local == str(path)
    assert info["xlsx_source_path"] == str(path)
    local, info = rules.resolve_xlsx("/Volumes/x/y/z.xlsx", None)
    assert local is None
    assert info["error"] == "no Files API client"


def test_fixture_main_reports_fetched_xlsx(monkeypatch, tmp_path: Path) -> None:
    content = _workbook_bytes()
    monkeypatch.setattr(run_recon, "XLSX_FETCH_OVERRIDE", lambda _path: content)
    reference = Path(__file__).parents[2] / "docs" / "migration" / "recon" / "reference"
    assert run_recon.main(
        [
            "--unit", "U4", "--mode", "fixture", "--business-date", "2024-01-31",
            "--out", str(tmp_path), "--target-dir", str(reference),
            "--xlsx-path", "/Volumes/reports/monthly.xlsx",
        ]
    ) == 0
    report = json.loads((tmp_path / "recon.json").read_text(encoding="utf-8"))
    assert report["xlsx"]["xlsx_sha256"] == hashlib.sha256(content).hexdigest()
    summary = (tmp_path / "recon.summary.md").read_text(encoding="utf-8")
    assert "xlsx: /Volumes/reports/monthly.xlsx sha256=" in summary
