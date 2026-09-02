from __future__ import annotations

import logging

import pytest
from openpyxl import load_workbook

from sas_macros import ParamError, export_xlsx, lock, nobs, sendmail, validate_param


def test_validate_param_required_missing() -> None:
    with pytest.raises(ParamError, match="Macro parameter REGION is required"):
        validate_param("region", None, required=True)


def test_validate_param_allowed_set() -> None:
    assert validate_param("region", "all", allowed=["ALL", "NE"]) == "ALL"
    with pytest.raises(ParamError, match="Allowable values are: ALL, NE"):
        validate_param("region", "SW", allowed=["ALL", "NE"])


@pytest.mark.parametrize(("value", "expected"), [("N", "0"), ("NO", "0"), ("Y", "1"), ("ON", "1")])
def test_validate_param_boolean_aliases(value: str, expected: str) -> None:
    assert validate_param("abort_on_err", value) == expected


def test_validate_param_numeric_constraints() -> None:
    assert validate_param("count", "3", allowed="POSITIVE") == "3"
    assert validate_param("count", "0", allowed="NONNEGATIVE") == "0"
    with pytest.raises(ParamError):
        validate_param("count", "0", allowed="POSITIVE")


def test_validate_param_multiword_rejection_and_default() -> None:
    with pytest.raises(ParamError, match="single word"):
        validate_param("name", "two words")
    assert validate_param("region", None, default="ALL") == "ALL"
    assert validate_param("description", "two words", words=True) == "TWO WORDS"


def test_nobs_with_fake_executor() -> None:
    class FakeExecutor:
        def execute(self, sql: str) -> list[tuple[object, ...]]:
            assert sql == "SELECT COUNT(*) FROM sas_legacy.sas_bronze.cust_accounts"
            return [(487,)]

    assert nobs(FakeExecutor(), "sas_legacy.sas_bronze.cust_accounts") == 487


def test_lock_returns_true() -> None:
    assert lock("sas_legacy.sas_bronze.cust_accounts") is True


def test_sendmail_returns_payload_and_logs(caplog: pytest.LogCaptureFixture) -> None:
    with caplog.at_level(logging.INFO):
        payload = sendmail("ops", "Subject", "Body", ["report.xlsx"])
    assert payload == {
        "to": "ops",
        "subject": "Subject",
        "body": "Body",
        "attachments": ("report.xlsx",),
    }
    assert "sendmail stub" in caplog.text


def test_export_xlsx_writes_header_and_rows(tmp_path) -> None:
    path = export_xlsx(
        [{"account_id": "A1", "balance": 12.5}, {"account_id": "A2", "balance": 8.0}],
        str(tmp_path / "nested" / "report.xlsx"),
    )
    workbook = load_workbook(path, read_only=True)
    rows = list(workbook.active.iter_rows(values_only=True))
    assert rows == [("account_id", "balance"), ("A1", 12.5), ("A2", 8.0)]
